#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Copyright (C) 2021-2022 Battelle Memorial Institute
# file: dsot_solar.py
# Created 2/27/2020
# @author: hard312 (Trevor Hardy)
"""This script provides a few essential functions:

- Download hourly annual solar data for a specified list of locations from the
  National Solar Radiation Database (NSRDB).

- Simulate the power production for a standard size solar array for each of the solar profiles
  downloaded using NREL's PySAM (Python wrapper for their SAM tool).

- Aggregate the individual solar profiles to form standardized distributed solar power profiles and
  centralized utility-scale profiles for the DSO+T 200-node model and 8-node model.

The script includes various housekeeping functions such as file
management, automatic generation of distributed solar sites around the
primary DSO location, and formatting of the necessary files for ingest by
their target simulators.

This script is structured like virtually all of my scripts are
structured: running this script with no arguments will perform a
comprehensive standard analysis using reasonable default inputs. All
necessary inputs and requested outputs will be produced in the "auto-run"
folder as a sort of combination example and test case.

Its worth noting that as of this writing, the NSRDB is VERY slow to
respond by the provided API and the daily limit of files requests is
relatively modest. To gather all the data to run ten distributed solar
sites per DSO (200-node) required literally a day.

Also note that use of this API requires registration and use of an NREL-
provided API key. This file has my API key removed so that I don't
inadvertently share its use with all of PNNL. Get your own key.

DSO+T is simulating 2016 which includes a Leap Day. The NSRDB data does
not include data for this day (which is odd, given that the data is based
on satellite photos; the source data should exist). I'm replicating data
from Feb 28th as the data for Feb. 29th, but I'm only doing this when
creating the output files. All the original NRSDB and SAM power profiles
only contains 365 days of data.
"""

import argparse
import datetime as dt
import json
import logging
import math
import os
import pprint
import random
import sys
import time
from pathlib import Path

import PySAM.Pvwattsv7 as pv
import matplotlib.pyplot as plt
import numpy as np
import openpyxl as xl
import pandas as pd
import requests

# Setting up logging
logger = logging.getLogger(__name__)

# Setting up pretty printing, mostly for debugging.
pp = pprint.PrettyPrinter(indent=4, )


def truncate(f, n=3):
    """
    This function truncates the passed in value to the specified
    number of decimals. The default value is three decimals.

    Args:
        f (string/float): Value to be truncated
        n (int): Number of decimal places to truncate to
    Returns:
        truc_val (string/float): Truncated value of f
    """
    s = '{}'.format(f)
    if 'e' in s or 'E' in s:
        return '{0:.{1}f}'.format(f, n)
    i, p, d = s.partition('.')
    trunc_val = '.'.join([i, (d + '0' * n)[:n]])
    return trunc_val


def _open_file(file_path, file_type='r'):
    """ Utilty function to open file with reasonable error handling.

    Args:
        file_path (str): Path to the file to be opened
        file_type (str): Type of the open method. Default is read ('r')
    Returns:
        fh (file object): File handle for the open file
    """
    try:
        fh = open(file_path, file_type)
    except IOError:
        logger.error('Unable to open {}'.format(file_path))
    else:
        return fh


def parse_solar_metadata(solar_metadata_path):
    """
    This function parses the solar metadata JSON.

    Args:
        solar_metadata_path (str): Path to the JSON file containing the metadata to be parsed.
    Returns:
        solar_dict (dicts): Dictionary form of the data in the JSON file
    """

    solar_fh = _open_file(solar_metadata_path)
    solar_dict = json.load(solar_fh)
    solar_fh.close()
    logger.info('Parsed solar metadata file {}'.format(solar_metadata_path))
    logger.info(pp.pformat(solar_dict))
    return solar_dict


def _create_200_node_csv_file_and_headers(filename, output_path):
    """ Writes out headers for solar data file where all solar data
    is stored as a table (rows for each timestamp, columns for each of
    the 200 buses).

        Args:
            filename (str): name of output file
            output_path (str): Path to location where output file will be written
        Returns:
            dsot_fh (File object): File handle for output file
        """
    outpath = os.path.join(output_path, filename)
    dsot_fh = _open_file(outpath, 'w')

    # Writing headers; every CSV needs headers, right?
    out_str = 'time,'
    num_dso = 200
    for idx in range(1, num_dso):
        out_str = out_str + 'solar' + str(idx) + ','

    # Last entry on each line is a bit special
    out_str = out_str + 'solar200\n'

    # Writing out headers to file early (mostly just to double-check during
    # development).
    dsot_fh.write(out_str)
    dsot_fh.flush()

    return dsot_fh


def _add_extra_days_to_hourly(profile):
    """
        This function adds extra buffer days to the beginning of an
        hourly profile and adds leap day to the profile.

        Checks are made on the incoming profile length to see what days
        need to be added.
            8760 - Add leap day and three warm-up days
            8784 - Assume leap day has already been added

        The added data is a replication of the first few days of Jan
        (for the days before Jan 1) and Feb 28th (for Leap Day).

        Args:
            profile (list): hourly profile values
        Returns:
            profile (list): Augmented hourly profile values
        """

    if len(profile) == 8760:
        add_buffer_days = True
        add_leap_day = True
        logger.info("\t...Adding buffer days to beginning of load profile...")
        logger.info("\t...Adding leap day to load profile...")
    elif len(profile) == 8784:
        add_buffer_days = True
        add_leap_day = False
        logger.info("\t...Adding buffer days to beginning of load profile...")
    elif len(profile) == 8856:
        add_buffer_days = False
        add_leap_day = False
        logger.info("\t...No additional days need to be added...")
    else:
        logger.error("Unsupported profile length.")
        logger.error(f"\tProfile length:{len(profile)}")
        raise SystemExit
    # Adjusting profile to add three buffer days at the beginning
    # (replication of Jan 1 data) and leap day (replication of
    # Feb 29th data).
    if add_leap_day:
        days_in_Jan = 31
        days = days_in_Jan + 28
        hours_per_day = 24
        hours = days * hours_per_day
        leap_day_profile = profile[hours - hours_per_day:hours]
        offset = hours
        for power in leap_day_profile:
            profile.insert(offset, power)
            offset = offset + 1
    # Adding three buffer days
    if add_buffer_days:
        prepend = profile[0:72]
        prepend.extend(profile)
        profile = prepend
    return profile


def parse_DSO_metadata_Excel(dso_metadata_path, worksheet_name):
    """
    (DEPRECATED): Metadata is now stored in JSON file. See parse_DSO_metadata_Excel_JSON.
    This function parses the DSO metadata which is contained in an Excel spreadsheet

    Args:
        dso_metadata_path (str): Path to the Excel file containing the metadata to be parsed.
        worksheet_name (str): Name of the worksheet in the Excel file containing the metadata
    Returns:
        dso_meta (list of dicts): One dictionary per DSO with appropriate metadata captured.
    """

    # openpyxl doesn't use file handles, so I just use my function to
    #   make sure the file is really there and then close it up to allo
    #   openpyxl to be able to grab it and do its thing.
    dso_fh = _open_file(dso_metadata_path)
    dso_fh.close()
    dso_meta = []
    wb = xl.load_workbook(dso_metadata_path)
    sheet = wb[worksheet_name]

    # Defining the location of the data I need to extract from each row
    #   of the spreadsheet
    header_idx = {}
    for header in sheet.iter_cols(min_row=1, max_row=1):
        if header[0].value == 'Latitude':
            header_idx['lat'] = header[0].col_idx
        if header[0].value == 'Longitude':
            header_idx['long'] = header[0].col_idx
        if header[0].value == 'Bus':
            header_idx['200-bus'] = header[0].col_idx
        if header[0].value == '8-Bus Mapping':
            header_idx['8-bus'] = header[0].col_idx
        if header[0].value == 'Average Load MW':
            header_idx['avg load'] = header[0].col_idx

    # Now iterate over the rest of the file to pull in the metadata into
    #   my dictionary
    for row in sheet.rows:
        # Skip the header row and the spacer row beneath
        if row[0].row > 2:
            for item in row:
                if item.col_idx == header_idx['lat']:
                    lat = item.value
                if item.col_idx == header_idx['long']:
                    long = item.value
                if item.col_idx == header_idx['200-bus']:
                    bus_200 = item.value
                if item.col_idx == header_idx['8-bus']:
                    bus_8 = item.value
                if item.col_idx == header_idx['avg load']:
                    avg_load = item.value
            dso_meta.append({'lat': lat,
                             'long': long,
                             '200-bus': bus_200,
                             '8-bus': bus_8,
                             'avg load': avg_load})
    logger.info('Parsed DSO metadata file {}'.format(dso_metadata_path))
    logger.info(pp.pformat(dso_meta))
    return dso_meta


def parse_DSO_metadata_Excel_JSON(dso_metadata_path_Excel, worksheet_name, dso_metadata_path_JSON):
    """
    This function parses the DSO metadata which is contained in a JSON and Excel files.
    Most of the metadata is in the JSON but one crucial piece of information is in the Excel file:
    the mapping of the 200-node to 8-node buses.

    Sample of the bus-generator file: (Note the first columns is empty)
    200 bus	8 bus
    1   1
    2   1
    3   1
    4   1
    5   1
    6   1
    ...

    Sample of JSON structure:
    "general": {
    ...
    }
    "DSO_1": {
    "bus_number": 1,
    "latitude": 33.02,
    "longitude": -96.85,
    "average_load_MW": 4154.30537,
    ...
    },
    ...

    Args:
        dso_metadata_path_Excel (str): Path to the Excel file containing the metadata to be parsed.
        worksheet_name (str): Name of the worksheet in the Excel file containing the metadata.
        dso_metadata_path_JSON (str): Path to the JSON file containing the metadata to be parsed.
    Returns:
        dso_meta (list of dicts): One dictionary per DSO with appropriate metadata captured.
    """
    # openpyxl doesn't use file handles, so I just use my function to
    #   make sure the file is really there and then close it up to allo
    #   openpyxl to be able to grab it and do its thing.
    dso_fh = _open_file(dso_metadata_path_Excel)
    dso_fh.close()
    dso_meta = []
    wb = xl.load_workbook(dso_metadata_path_Excel)
    sheet = wb[worksheet_name]

    # Defining the location of the data I need to extract from each row
    #   of the spreadsheet
    header_idx = {}
    for header in sheet.iter_cols(min_row=1, max_row=1):
        if header[0].value == '200 bus':
            header_idx['200-bus'] = header[0].col_idx
        if header[0].value == '8 bus':
            header_idx['8-bus'] = header[0].col_idx

    # Now iterate over the rest of the file to pull in the metadata into
    #   my dictionary
    for row in sheet.rows:
        # Skip the header row and the spacer row beneath
        if row[0].row > 1:
            for item in row:
                if item.col_idx == header_idx['200-bus']:
                    bus_200 = item.value
                if item.col_idx == header_idx['8-bus']:
                    bus_8 = item.value
            dso_meta.append({'200-bus': bus_200, '8-bus': bus_8})
    logger.info('Parsed DSO Excel metadata file {}'.format(dso_metadata_path_Excel))
    logger.info(pp.pformat(dso_meta))

    # Adding in the rest of the metadata from the JSON file.
    fh = _open_file(dso_metadata_path_JSON)
    json_meta = json.load(fh)
    fh.close()

    for idx, dso in enumerate(dso_meta):
        dso_num = idx + 1
        dso['lat'] = json_meta[f'DSO_{dso_num}']['latitude']
        dso['long'] = json_meta[f'DSO_{dso_num}']['longitude']
        dso['avg load'] = json_meta[f'DSO_{dso_num}']['average_load_MW']

    logger.info('Parsed JSON Excel metadata file {}'.format(dso_metadata_path_Excel))
    log_metdata(dso_meta)

    return dso_meta


def build_dso_solar_folders(dso_meta, output_path):
    """
    Creates folders for the solar data in the output path. Folders are
    only created if they don't already exist.

    Args:
        dso_meta: metadata dictionary containing the DSO name
        output_path: string specifying output path of data
    """
    for dso in dso_meta:
        dir_name = 'DSO_' + str(dso['200-bus'])
        dir_path = Path(output_path, dir_name)
        dir = Path(dir_path)
        if dir.is_dir():
            pass
        else:
            os.mkdir(dir_path)


def add_locations(dso_meta, solar_meta, nsrdb_path):
    """
    This function adds a number of random locations (specified in
    solar_meta) withing a certain distance (specified in solar_meta) of
    a DSO to replicate the effects of the distributed solar generation.
    Also randomizes the tilt and azimuth orientation of the panels for
    each distributed location.

    The primary location for the DSO is the lat/long from the DSO
    metadata file and is also added to the solar site list. This is the
    location that will be used for any utility-scale solar generation.

    Args:
        dso_meta (list): List of dicts with metadata associated with each DSO
        solar_meta (dict): Metadata related to the solar parameters
        nsrdb_path(str): Path to the directory with the NSRDB data
    Returns:
        dso_meta (list of dicts): One dictionary per DSO with appropriate
        metadata captured. Each dictionary has an added item,
        "distributed locations" which is a dict of parameters for the
        distributed solar generation
    """

    count = solar_meta['distributed_sites']['count']
    km_per_lat_deg = 110.574

    for idx, dso in enumerate(dso_meta):
        dso['solar_sites'] = []
        dir_name = 'DSO_' + str(dso['200-bus'])
        location_file_name = 'DSO_' + str(dso['200-bus']) + '_locations.json'
        file_path = Path(nsrdb_path, dir_name, location_file_name)

        # Opening existing location file (if there is one)
        file = Path(file_path)
        if file.is_file():
            loc_fh = _open_file(file)
            dso['solar_sites'] = json.load(loc_fh)
            loc_fh.close()
            logger.info('Loaded in location list for DSO {}'.format(
                dso['200-bus']))
        else:
            # Adding in primary DSO location
            # Fix azimuth and tilt to semi-optimal values for primary location
            dso['solar_sites'].append({'lat': dso['lat'], 'long': dso['long'],
                                       'azimuth': 180, 'tilt': 30, 'downloaded': 'no'})
            logger.info('Added DSO primary location at {}, {} to solar site '
                        'list'.format(dso['lat'], dso['long']))

        # Checking to see how many points are in the location file and
        #   checking to see if we need to add more. Checking DSO on first
        #   transmission node arbitrarily; all DSOs will have the same number
        #   of locations.
        #
        # There should always be at least one lcoation, the primary DSO
        #   location that was just added if the location file was empty.
        dso_site_count = len(dso['solar_sites']) - 1

        if count > dso_site_count:
            new_site_count = count - dso_site_count

            # Ensuring that when we add new locations we use a different seed
            #   such that we don't always start with the same seed and thus
            #   add the same locations when adding new locations
            random.seed(dso_site_count)

            base_lat = dso['lat']
            base_long = dso['long']
            km_per_long_deg = 111.320 * math.cos(math.radians(base_lat))
            lat_offset = solar_meta['distributed_sites']['offset_km'] / km_per_lat_deg
            long_offset = solar_meta['distributed_sites'][
                              'offset_km'] / km_per_long_deg
            # Assumes we're far enough away from the equator that we won't
            #   have to worry about negative values
            min_lat = base_lat - lat_offset
            max_lat = base_lat + lat_offset

            # Assumes we're far enough away from the prime meridian that we
            #   won't have to worry about negative values
            min_long = base_long - long_offset
            max_long = base_long + long_offset

            for site in range(0, new_site_count):
                lat = random.uniform(min_lat, max_lat)
                long = random.uniform(min_long, max_long)
                tilt = random.uniform(
                    solar_meta['rooftop_tilt_limits']['lower_limit_deg'],
                    solar_meta['rooftop_tilt_limits']['upper_limit_deg'])
                # Assumes two values 'south' and 'west'
                # Typically south will be larger than west, but I'm going to be
                #   thorough and account for cases when it is not.
                if solar_meta['azimuth_factor']['south'] > solar_meta[
                    'azimuth_factor']['west']:
                    if random.random() > solar_meta['azimuth_factor']['west']:
                        azimuth = 180  # southerly
                    else:
                        azimuth = 270  # westerly
                else:  # when west is a greater fraction than south
                    if random.random() > solar_meta['azimuth_factor']['south']:
                        azimuth = 270
                    else:
                        azimuth = 180

                site_data = {'lat': lat,
                             'long': long,
                             'tilt': tilt,
                             'azimuth': azimuth,
                             'downloaded': 'no'}
                dso['solar_sites'].append(site_data)
        # Saving out the additions to the metadata in the original data
        #   structure
        logger.info('Added random solar site locations')
        logger.info(pp.pformat(dso['solar_sites']))
        dso_meta[idx] = dso

        # Once we've created the site list dump it out to file
        json_fh = open(file_path, 'w')
        json.dump(dso['solar_sites'], json_fh)
        json_fh.close()
        logger.info('Wrote solar site location data to file: {}'.format(
            file_path))

    return dso_meta


def generate_KML(dso_meta, output_file):
    """
    This function attempts to make a KML file that can be loaded
    into Google Earth to visualize the solar locations.

    Args:
        dso_meta (list): List of dicts with metadata associated with each DSO
        output_file (str): Location to write output KML
    """

    kml_fh = open(output_file, 'w')

    # Write headers
    kml_fh.write('<?xml version="1.0" encoding="UTF-8"?>\n')
    kml_fh.write('<kml xmlns="http://earth.google.com/kml/2.0">\n')
    kml_fh.write('<Document>\n')

    # Build icon style URLs
    icon_url_head = 'http://maps.google.com/mapfiles/kml/paddle'
    icon_color_list = [
        'blu',
        'grn',
        'ltblu',
        'pink',
        'purple',
        'red',
        'wht',
        'ylw',
        'orange'
    ]
    icon_shape_list = [
        'blank',
        'circle',
        'diamond',
        'square',
        'stars'
    ]
    icon_str_list = []
    for color in icon_color_list:
        for shape in icon_shape_list:
            icon_str_list.append('{}/{}-{}.png'.format(icon_url_head, color,
                                                       shape))
    # Duplicating list, so I have enough icons for all 200 DSOs
    icon_str_list = icon_str_list + icon_str_list + icon_str_list + \
                    icon_str_list + icon_str_list

    # Writing style definitions for the markers
    for idx, dso in enumerate(dso_meta):
        # Writing style information
        id = 'DSO_' + str(dso['200-bus'])
        kml_fh.write('<Style id="{}">\n'.format(id))
        kml_fh.write('\t<IconStyle>\n')
        kml_fh.write('\t\t<scale>0.75</scale>\n')
        kml_fh.write('\t\t<Icon>\n')
        kml_fh.write('\t\t\t<href>{}</href>\n'.format(icon_str_list[idx]))
        kml_fh.write('\t\t</Icon>\n')
        kml_fh.write('\t</IconStyle>\n')
        kml_fh.write('</Style>\n')

        # Writing marker locations
        for site in dso['solar_sites']:
            kml_fh.write('<Placemark>\n')
            kml_fh.write('\t<styleUrl>#{}</styleUrl>\n'.format(id))
            kml_fh.write('\t<Point><coordinates>{},{},'
                         '0</coordinates></Point>\n'.format(site['nsrdb_long'],
                                                            site['nsrdb_lat']))
            kml_fh.write('</Placemark>\n')
    kml_fh.write('</Document>\n')
    kml_fh.write('</kml>')
    kml_fh.close()

    logger.info('Wrote out KML of all solar sites to file {}'.format(
        output_file))


def download_nsrdb_data(dso_meta, solar_meta, output_path):
    """
    This function queries the NSRDB database over the web and pulls
    down the solar data down and stores it in a Pandas dataframe.

    Args:
        dso_meta (list): List of dicts with metadata associated with
            each DSO, specifically the site information
        solar_meta (list): List of metadata related to the solar parameters
        output_path (str): Location to write out solar data from the NSRDB
            so that we don't have to re-query the database for data we already have.
    Returns:
        dso_meta (list): List of dicts with metadata associated with
        each DSO, updated to include NSRDB-specific information such as
        the abreviated lats and longs from the downloaded solar data as
        well as a path to each downloaded file and a boolean indicating
        a file has been downloaded for the given site.
    """

    # Declare all variables as strings. Spaces must be replaced with '+', i.e., change 'John Smith' to 'John+Smith'.
    # Define the lat, long of the location and the year
    for idx, dso in enumerate(dso_meta):
        dist_count = solar_meta['distributed_sites']['count']
        total_count = dist_count + 1  # add primary DSO site
        dso_dir = 'DSO_' + str(dso['200-bus'])
        outpath = os.path.join(output_path, dso_dir)
        file_list = os.listdir(outpath)

        # Assume all CSVs are solar data files.
        # This techinque has a problem maintaining a consistent state
        #   between what the JSON shows and what files have been
        #   downloaded. Rather than just counting files with the
        #   correct extension I should do a comparison between the files
        #   that I've downloaded and what the JSON says, getting the
        #   JSON to match the state of the files. Since this is not
        #   something I have to run often, for now it just makes sense
        #   to do things this way. If a particular DSO gets into an
        #   inconsistent state I can just blow away all the files and
        #   rebuild that particular solar data list from scratch.
        # Doing the quick and dirty for thing for now...
        downloaded_count = 0
        for file in file_list:
            name, ext = os.path.splitext(file)
            if ext == '.csv':
                downloaded_count = downloaded_count + 1
        files_to_download = total_count - downloaded_count
        if files_to_download > 0:
            for idx2, site in enumerate(dso['solar_sites']):
                if site['downloaded'] == 'no' and files_to_download > 0:
                    # TDH: largely copied from
                    #   https://nsrdb.nrel.gov/data-sets/api-instructions.html
                    lat = site['lat']
                    long = site['long']
                    year = str(2016)
                    logger.info('Querying NSRDB for data at site {}, {} '
                                ''.format(lat, long))

                    # You must request an NSRDB api key from https://developer.nrel.gov/signup/
                    api_key = 'put NSRDB API key here as a string'
                    # Set the attributes to extract (e.g., dhi, ghi, etc.), separated by commas.
                    attributes = 'ghi,dhi,dni,wind_speed,air_temperature,solar_zenith_angle'
                    # Set leap year to true or false. True will return leap day data if present, false will not.
                    leap_year = 'false'
                    # Set time interval in minutes, i.e., '30' is half hour intervals. Valid intervals are 30 & 60.
                    interval = '60'
                    # Specify Coordinated Universal Time (UTC), 'true' will use UTC, 'false' will use the local time zone of the data.
                    # NOTE: In order to use the NSRDB data in SAM, you must specify UTC as 'false'. SAM requires the data to be in the
                    # local time zone.
                    utc = 'false'
                    # Your full name, use '+' instead of spaces.
                    your_name = 'Trevor+Hardy'
                    # Your reason for using the NSRDB.
                    reason_for_use = 'research'
                    # Your affiliation
                    your_affiliation = 'PNNL'
                    # Your email address
                    your_email = 'trevor.hardy@pnnl.gov'
                    # Please join our mailing list, so we can keep you up-to-date on new developments.
                    mailing_list = 'false'

                    # Declare url string
                    url = 'https://developer.nrel.gov/api/solar/nsrdb_psm3_download.csv?wkt=POINT({lon}%20{lat})&names={year}&leap_day={leap}&interval={interval}&utc={utc}&full_name={name}&email={email}&affiliation={affiliation}&mailing_list={mailing_list}&reason={reason}&api_key={api}&attributes={attr}'.format(
                        year=year, lat=lat, lon=long, leap=leap_year,
                        interval=interval, utc=utc, name=your_name,
                        email=your_email, mailing_list=mailing_list,
                        affiliation=your_affiliation, reason=reason_for_use,
                        api=api_key, attr=attributes)
                    # Return just the first 2 lines to get metadata:
                    logger.info('\tGetting headers for NSRDB data at {}, {}'
                                ''.format(lat, long))
                    info = pd.read_csv(url, nrows=1)
                    # info = []
                    # r = requests.request("POST", f'{url}&{info}')
                    # See metadata for specified properties, e.g., timezone and elevation
                    timezone, elevation = info['Local Time Zone'], info['Elevation']
                    nsrdb_lat, nsrdb_long = info['Latitude'], info['Longitude']
                    nsrdb_lat = float(nsrdb_lat)
                    nsrdb_long = float(nsrdb_long)

                    # Check to see if file exists (indicating we downloaded it before
                    #   and don't need to do so again). If file does exist, we load
                    #   it and add it to the list of dataframes.
                    filename = '{}_{}_solar_data.csv'.format(nsrdb_lat,
                                                             nsrdb_long)
                    output_file = os.path.join(output_path, dso_dir, filename)

                    r = requests.get(url, allow_redirects=True)
                    csv_fh = open(output_file, 'wb')
                    csv_fh.write(r.content)
                    csv_fh.close()

                    # Saving back to original data structure
                    site['downloaded'] = 'yes'
                    site['nsrdb_lat'] = nsrdb_lat
                    site['nsrdb_long'] = nsrdb_long
                    site['nsrdb_file'] = output_file
                    dso['solar_sites'][idx2] = site
                    files_to_download = files_to_download - 1

                    logger.info('\tDownloaded and saved data for {}, '
                                '{}'.format(
                        site['nsrdb_lat'], site['nsrdb_long']))
                    # Taking a break to not overload the NSRDB server
                    time.sleep(5)

            # Saving back to original data strcture
        dso_meta[idx] = dso
        filename = dso_dir + '_locations.json'
        file_path = os.path.join(output_path, dso_dir, filename)
        json_fh = open(file_path, 'w')
        json.dump(dso['solar_sites'], json_fh)
        json_fh.close()
        logger.info('All necessary solar data collected, saved solar data '
                    'status for DSO {} in file: {}'.format(idx + 1, file_path))
    logger.info('Loaded all NSRDB data for this run.')
    return dso_meta


def calc_solarPV_power(dso_meta, output_path):
    """
    This function uses PySAM from NREL
    (https://sam.nrel.gov/software-development-kit-sdk/pysam.html)
    to calculate the solar PV power generation for a solar array at the
    indicated site. Metadata associated with each site determines the
    tilt and azimuth. All other array modeling parameters (largely
    losses) are constant across all sites. The array size is also fixed
    at a 1 MW.

    Args:
        dso_meta (list): List of dicts with metadata associated with each DSO,
            specifically the solar PV information per site
        output_path (str): Location to write out solar PV power data
    Returns:
        dso_meta (list): List of dicts with metadata associated with
        each DSO, updated to include power profiles for each site in the DSO.
    """
    for idx, dso in enumerate(dso_meta):
        power_profiles = []
        logger.info('Processing solar profiles for DSO {}'.format(
            dso['200-bus']))
        for idx2, site in enumerate(dso['solar_sites']):
            filename = '{}_{}_1MW_hourly_annual_power_profile.csv'.format(
                site['nsrdb_lat'], site['nsrdb_long'])
            dso_dir = 'DSO_' + str(dso['200-bus'])
            outpath = os.path.join(output_path, dso_dir, filename)
            file = Path(outpath)
            if file.is_file():
                logger.info('\tSolar profile already exists and not '
                            'recreated for {}, {}'
                            ''.format(site['nsrdb_lat'],
                                      site['nsrdb_long']))
                # Read in power profile for later use
                power_list = []
                power_fh = _open_file(outpath)
                power = power_fh.readlines()
                power = [x.strip() for x in power]
                power = [float(x) for x in power]
                power_profiles.append(power)

            else:
                logger.info('\t\tCreating solar profile for {}, {}'.format(
                    site['nsrdb_lat'], site['nsrdb_long']))
                pv_model = pv.default('PVWattsResidential')
                pv_model.SolarResource.solar_resource_file = site['nsrdb_file']

                # Setting array parameters
                # Fixed rack
                pv_model.SystemDesign.array_type = 0
                pv_model.SystemDesign.tilt = site['tilt']
                pv_model.SystemDesign.azimuth = site['azimuth']

                # Using all defaults from SAM GUI
                pv_model.SystemDesign.dc_ac_ratio = 1.2
                # Sum of shading, mismatch, wiring, connections,
                # soiling and nameplate
                pv_model.SystemDesign.losses = 12  # percent

                pv_model.SystemDesign.system_capacity = 1000  # kW
                pv_model.execute()
                power_profile = pv_model.Outputs.ac  # Output profile in W
                power_profile = [x / 1000000 for x in power_profile]  # MW

                # Saving power profile in memory for ease of later manipulation
                power_profiles.append(power_profile)
                filename = '{}_{}_1MW_hourly_annual_power_profile.csv'.format(
                    site['nsrdb_lat'], site['nsrdb_long'])
                outpath = os.path.join(output_path, dso_dir, filename)
                write_power_profile(outpath, power_profile)
        logger.info('Completed all solar profile power calculation for DSO {}\n'.format(dso['200-bus']))
        dso['power_profiles'] = power_profiles
        dso_meta[idx] = dso
    # TDH: My four-year-old son wants you to know he can type his own name:
    #  avery
    return dso_meta


def write_power_profile(output_path, power_data):
    """
    This function writes out a power profile to the provided file.
    Only intended to write out a single value per line. I probably should
    have used the standard CSV library instead.

    Args:
        power_data (list): List of power values
        output_path (str): Path to the location of file to write.
    """
    power_fh = open(output_path, 'w')
    for power in power_data:
        power_fh.write(str(truncate(power)) + '\n')
    power_fh.close()


def calc_dso_solar_fraction(dso_meta):
    """
    This function calculates the total target PV fraction in
    proportion with the total average load of the system. That is, it
    answers the question, "What fraction of the total installed PV
    should be allocated to each DSO?" (The total size of the installed
    PV for all of ERCOT is pre-defined by another analysis.)

    Args:
        dso_meta (list): List of dicts with metadata associated with each DSO,
            specifically the average annual load of each DSO
    Returns:
        dso_meta (list): List of dicts with metadata associated with
        each DSO, updated to include load fractions for each DSO.
    """
    total_load = 0
    for idx, dso in enumerate(dso_meta):
        total_load = total_load + dso['avg load']
    for dso in dso_meta:
        dso['load_fraction'] = dso['avg load'] / total_load
        dso_meta[idx] = dso
    return dso_meta


def aggregate_scale_solar_pv_profiles(dso_meta, solar_meta, output_path):
    """
    This function calculates the utility-scale and distributed solar
    PV power profiles. The utility scale profile is from a single location
    (the primary DSO location) while the distributed profile is an
    equally-weighted average of all profiles for the DSO.

    Both the utility-scale and distributed power profiles are scaled to
    meet both the total solar capacity contribution of the DSO and
    their relative ratios between utility-scale and dsitributed.

    Args:
        dso_meta (list): List of dicts with metadata associated with each DSO,
            pecifically with the load fraction needed to scale
            the solar PV power profiles as well as the profiles themselves
        solar_meta (dict): User-specified solar metadata
        output_path (str): Path to the solar PV power profiles for all DSOs.
    Returns:
        dso_meta (list): List of dicts with metadata associated with
        each DSO, updated to include the scaled and aggregated
        utility-scale and dsitributed solar profiles.
    """
    head, tail = os.path.split(output_path)
    filename = 'DSOT solar PV capacities.csv'
    outpath = os.path.join(head, filename)
    dso_cap_fh = _open_file(outpath, 'w')

    # Writing out headers
    dso_cap_fh.write('DSO number,Utility-scale installed capacity (MW),'
                     'Distributed-capacity (MW)\n')

    aggreagte_dist_solar = []
    for idx, dso in enumerate(dso_meta):
        logger.info('Calculating and applying scaling factors for DSO {}'.format(dso['200-bus']))
        # Initializing total power profile
        agg_power_profile = [0] * 8760
        for power in dso['power_profiles']:
            agg_power_profile = [x + y for x, y in zip(agg_power_profile,
                                                       power)]
        dso['agg_power_profile'] = agg_power_profile

        # Distributed solar
        # One MW per solar profile
        dso['summed_profile_capacity_MW'] = len(dso['power_profiles'])
        target_total_solar = dso['load_fraction'] * solar_meta[
            'total_solar_installation_capacity_MW']
        dso['total solar capacity'] = target_total_solar
        target_dist_solar = target_total_solar * solar_meta[
            'ownership_factor']['customer_owned']
        dso['dist solar capacity'] = target_dist_solar
        dist_scaling_factor = target_dist_solar / dso[
            'summed_profile_capacity_MW']
        agg_power_profile = [x * dist_scaling_factor for x in
                             agg_power_profile]
        dso['dist power profile'] = agg_power_profile
        dso_dir = 'DSO_' + str(dso['200-bus'])
        filename = 'DSO_{}_distributed_hourly_annual_power_profile.csv'.format(
            dso['200-bus'])
        outpath = os.path.join(output_path, dso_dir, filename)
        write_power_profile(outpath, agg_power_profile)
        dso['dist power profile path'] = outpath
        logger.info('\tDistributed scaling factor: {}'.format(
            dist_scaling_factor))

        # Utility-scale solar
        target_utility_solar = target_total_solar * solar_meta[
            'ownership_factor']['utility_owned']
        dso['utility solar capacity'] = target_utility_solar

        # Writing out scaling factor for use by others
        dso_cap_fh.write(str(dso['200-bus']) + ',' + str(target_utility_solar) +
                         ',' + str(target_dist_solar) + '\n')

        # Each profile is scaled to 1 MW and the utility-scale solar
        #   has only one profile
        utility_scaling_factor = target_utility_solar / 1
        # The first power profile is for the primary DSO site which is
        #   used as the site for all utility-scale solar PV.
        agg_power_profile = [x * utility_scaling_factor for x in
                             dso['power_profiles'][0]]
        dso['utility power profile'] = agg_power_profile
        filename = 'DSO_{}_utility_hourly_annual_power_profile.csv' \
                   ''.format(dso['200-bus'])
        outpath = os.path.join(output_path, dso_dir, filename)
        write_power_profile(outpath, agg_power_profile)
        dso['utility power profile path'] = outpath
        logger.info('\tUtility scaling factor: {}'.format(
            utility_scaling_factor))

        # Sum utilty-scale DSO solar profile profile to create 8-node
        #   profile

        dso_meta[idx] = dso
    log_metdata(dso_meta)
    return dso_meta


def aggregate_to_8_nodes(dso_meta, output_path):
    """
    This function aggregates the individual profiles for each of the
    200 nodes into a single profile for each of the reduced 8 nodes in
    the smaller model. This is done for both the utility and the
    distributed solar profiles.

    The 8-node DSOs are added to the dso_meta object to the end of the
    list. They don't have all the same metadata as the original 200-node
    DSOs.

    Args:
        dso_meta (list): List of dicts with metadata associated with each DSO,
            specifically 200-node DSO solar profiles
        output_path (str): Location to write out aggregated solar
        profile data.
    Returns:
        dso_meta (list): List of dicts with metadata associated with
        each DSO, updated to include aggregated solar profile data and
        output file location.
    """
    # Creating 8-node folders in output path
    for eight_node_dso in range(1, 9):
        # The 8-node DSOs will have the indices 200-207
        dso_meta.append({})
        dir_name = '8-node DSO_' + str(eight_node_dso)
        dir_path = Path(output_path, dir_name)
        dir = Path(dir_path)
        if dir.is_dir():
            pass
        else:
            os.mkdir(dir_path)
        dso_meta[200 + eight_node_dso - 1]['8-node dso num'] = eight_node_dso
        dso_meta[200 + eight_node_dso - 1]['dist power profile'] = \
            [0] * 8760
        dso_meta[200 + eight_node_dso - 1]['utility power profile'] = \
            [0] * 8760
        dso_meta[200 + eight_node_dso - 1]['dist solar capacity'] = 0
        dso_meta[200 + eight_node_dso - 1]['utility solar capacity'] \
            = 0
        dso_meta[200 + eight_node_dso - 1]['total solar capacity'] \
            = 0
        dso_meta[200 + eight_node_dso - 1]['avg load'] = 0
        logger.info('Initialized 8-node DSO {}'.format(eight_node_dso))

    # Summing the distributed solar PV power profiles up into 8 profiles
    for idx, dso in enumerate(dso_meta):
        # Last 8 DSOs in list are the 8-node DSOs and don't have power
        #   profiles to add to the aggregated 8-node data
        if idx < 200:
            eight_node_idx = 200 + dso['8-bus'] - 1
            logger.info('Added solar PV power profile from 200 node DSO {'
                        '}...'.format(idx))
            # Summing up distributed power profiles to create single,
            #   aggregated profile for 8-node DSO
            dso_meta[eight_node_idx]['dist power profile'] = [
                x + y for x, y in zip(
                    dso_meta[eight_node_idx]['dist power profile'],
                    dso['dist power profile'])]

            # Summing up utility power profiles to create single,
            #   aggregated profile for 8-node DSO
            dso_meta[eight_node_idx]['utility power profile'] = [
                x + y for x, y in zip(
                    dso_meta[eight_node_idx]['utility power profile'],
                    dso['utility power profile'])]

            # Summing capacity values for reporting purposes
            dso_meta[eight_node_idx]['dist solar capacity'] = \
                dso_meta[eight_node_idx]['dist solar capacity'] + \
                dso['dist solar capacity']
            dso_meta[eight_node_idx]['utility solar capacity'] = \
                dso_meta[eight_node_idx]['utility solar capacity'] \
                + dso['utility solar capacity']
            dso_meta[eight_node_idx]['total solar capacity'] = \
                dso_meta[eight_node_idx]['total solar capacity'] \
                + dso['total solar capacity']
            dso_meta[eight_node_idx]['avg load'] = \
                dso_meta[eight_node_idx]['avg load'] + dso['avg load']
            logger.info('\t... to 8-node DSO {}...'.format(dso['8-bus']))

    # Saving out results
    for eight_node_dso in range(1, 9):
        dir_name = '8-node DSO_' + str(eight_node_dso)
        filename = '8-node DSO {} dist solar PV power profile.csv'.format(
            eight_node_dso)
        outpath = os.path.join(output_path, dir_name, filename)
        dso_meta[200 + eight_node_dso - 1]['dist power profile path'] \
            = outpath
        write_power_profile(
            outpath,
            dso_meta[200 + eight_node_dso - 1]['dist power profile'])

        filename = '8-node DSO {} utility solar PV power profile.csv'.format(
            eight_node_dso)
        outpath = os.path.join(output_path, dir_name, filename)
        dso_meta[200 + eight_node_dso - 1]['utility power profile ' \
                                           'path'] = outpath
        write_power_profile(outpath, dso_meta[200 + eight_node_dso - 1][
            'utility power profile'])

        logger.info('Saving output power profiles to {}'.format(
            outpath))
        logger.info('DSO {} 8-node solar capacities:'.format(eight_node_dso))
        logger.info('\t Average Load (MW): {}'.format(dso_meta[200 +
                                                               eight_node_dso - 1]['avg load']))
        logger.info('\t Total solar capacity (MW): {}'.format(dso_meta[200 +
                                                                       eight_node_dso - 1]['total solar capacity']))
        logger.info('\t Utility solar capacity (MW): {}'.format(dso_meta[200 +
                                                                         eight_node_dso - 1]['utility solar capacity']))
        logger.info('\t Distributed solar capacity (MW): {}'.format(dso_meta[
                                                                        200 + eight_node_dso - 1][
                                                                        'dist solar capacity']))

    return dso_meta


def _calc_hours(month, start_day, num_days):
    """
    This function calculates the starting and ending hours for a
    user-provided date range. The data being graphed is hourly and can
    be accessed by index.

    Args:
        month (int): 1 to 12, indicating which month of the year to graph
        start_day (int): 1 to 31 (depending on the month) day of the month to start graphing
        num_days (int): Number of days in data to graph.
    Returns:
        first_hour (int): First hour (index) to be graphed
        last_hour (int): Last hour (index) to be graphed
    """
    days_in_months = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    first_hour = sum(days_in_months[:month]) * 24
    last_hour = first_hour + (num_days * 24)
    return first_hour, last_hour


def create_graphs(dso_meta, type):
    """
    This function graphs four weeks out of the year for each DSO to
    enable easy comparison between the effects of the distributed and
    utility (single-point) solar power profiles.

    Graph image files are saved alongside solar profile files.

    Args:
        dso_meta (list): List of dicts with metadata associated with each DSO,
            specifically the distributed and utility scale solar power profiles.
        type (str): Keyword indicating the type of graph to create
    """
    timestamps = np.arange('2016-01-01 00:00', '2016-12-31 23:00',
                           dtype='datetime64[h]')
    for idx, dso in enumerate(dso_meta):
        if idx < 200:
            logger.info('Collecting data to graph for DSO {}...'.format(dso[
                                                                            '200-bus']))
        else:
            logger.info('Collecting data to graph for 8-node DSO {}...'.format(
                dso['8-node dso num']))
        month_list = [1, 4, 7, 10]
        month_name_list = ['January', 'April', 'July', 'October']
        for idx2, month in enumerate(month_list):
            hour_start, hour_end = _calc_hours(month, 14, 7)
            ts = timestamps[hour_start:hour_end]
            fig = plt.figure(figsize=(12, 8))
            if type == 'profiles':
                dist_data = dso['dist power profile'][hour_start:hour_end]
                utility_data = dso['utility power profile'][hour_start:hour_end]

                plt.plot(ts, dist_data)
                plt.plot(ts, utility_data)
                plt.legend(['distributed solar PV', 'utility solar PV'])
                if idx < 200:
                    plt.title('DSO {} {} Solar PV Power Profile Comparison'.format(
                        dso['200-bus'], month_name_list[idx2]))
                else:
                    plt.title('DSO {} {} Solar PV Power Profile Comparison'
                              ''.format(['8-node dso num'], month_name_list[idx2]))
                plt.xticks(rotation=90)
                plt.xlabel('Date')
                plt.ylabel('Output Power (MW)')
                if idx < 200:
                    head, tail = os.path.split(dso['utility power profile path'])
                    filename = os.path.join(head, 'DSO {} {} Solar PV Profile'
                                                  ''.format(dso['200-bus'], month_name_list[idx2]))
                else:
                    head, tail = os.path.split(dso['dist power profile '
                                                   'path'])
                    filename = os.path.join(head, '8-node DSO {} {} Solar PV Profile'
                                                  ''.format(['8-node dso num'], month_name_list[idx2]))
                plt.savefig(filename, bbox_inches='tight')
            elif type == 'utility forecast':
                utility_data = dso['utility power profile'][
                               hour_start:hour_end]
                forecast_data = dso['utility forecast profile'][
                                hour_start:hour_end]

                plt.plot(ts, forecast_data)
                plt.plot(ts, utility_data)
                plt.legend(['Utility Forecast', 'Utility Actual'])
                if idx < 200:
                    plt.title(
                        'DSO {} {} Solar PV Power Forecast Comparison'.format(
                            dso['200-bus'], month_name_list[idx2]))
                else:
                    plt.title('DSO {} {} Solar PV Power Forecast Comparison'
                              ''.format(['8-node dso num'],
                                        month_name_list[idx2]))
                plt.xticks(rotation=90)
                plt.xlabel('Date')
                plt.ylabel('Output Power (MW)')
                if idx < 200:
                    head, tail = os.path.split(
                        dso['utility power profile path'])
                    filename = os.path.join(head, 'DSO {} {} Solar PV '
                                                  'Forecast Comparison'.format(
                        dso['200-bus'], month_name_list[idx2]))
                else:
                    head, tail = os.path.split(dso['dist power profile '
                                                   'path'])
                    filename = os.path.join(head,
                                            '8-node DSO {} {} Solar PV '
                                            'Forecast Comparison'.format(
                                                dso['8-node dso num'],
                                                month_name_list[idx2]))
                plt.savefig(filename, bbox_inches='tight')
            plt.close(fig)
        logger.info('\t... and saving plots in {}'.format(head))


def forecast_cleanup(dso_meta, idx, error, profile, forecast_profile):
    """
    This function cleans up the forecast power profiles to eliminate
    negative forecast values and non-zero forecasts when the actual
    values are zero (used as an indication that it is night). Not a
    great approximation but good enough.

    The 8-node DSOs are added to the dso_meta object to the end of the
    list. They don't have all the same metadata as the original 200-node
    DSOs.

    Args:
        dso_meta (list): List of dicts with metadata associated with each DSO
        idx (int): Index for dso_meta to indicate which DSO needs to be cleaned up.
        error (list): List of values containing the synthesized error signal
        forecast_profile (list): List of values containing the forecast
            profile with forecast errors added (sometimes causing problems
            that this function cleans up).
    Returns:
        forecast_profile (list): List of values containing the forecast
        profile after cleaning.
    """

    # I'm sure there's a list comprehension way to do this but
    #   this is what I'm going to do for now.
    # When actual solar is zero assume the sun is below the
    #   horizon and forecast would be zero. Also removing
    #   negative forecast values.
    for idx2, val in enumerate(error):
        correction = 0
        if profile[idx2] == 0:
            forecast_profile[idx2] = 0
        if forecast_profile[idx2] < 0:
            forecast_profile[idx2] = 0
    return forecast_profile


def create_hourly_solar_forecast(dso_meta, dso_type, rng_seed):
    """
    This function creates an hourly forecast for the utility profile
    by adding noise from a random distribution to it. The parameters for
    said distribution come from literature
    (https://www.nrel.gov/docs/fy15osti/63876.pdf).

    In a clumsy way that I'm not proud of now that I'm documenting the
    function, the user decides whether to add noise to the 8-node DSOs
    or the 200-node DSOs by defining a value for the "dso_type"
    parameter.

    Args:
        dso_meta (list): List of dicts with metadata associated with
            each DSO including the current utility dso power profiles.
        dso_type (int): Used to indicate whether to create forecast
            files for the 8-node or 200-node utility power profiles.
            Valid values are "8" or "200".
    Returns:
        dso-Meta (list): List of dicts with DSO data including the
        synthesized forecast profiles.
    """

    if dso_type == 200:
        dso_idxs = range(0, 200)
        np.random.seed(rng_seed + 200)
    else:
        dso_idxs = range(200, 208)
        np.random.seed(rng_seed + 8)

    lit_installed_capacity = 4173  # MW
    lit_std_dev = 168

    normalized_std_dev = lit_std_dev / lit_installed_capacity

    # Since some solar production is zero (during the night), I assume
    #   that the forecasts during these periods will be very accurate. The
    #   "forecast_cleanup" function goes through and removes the forecasts that
    #   are negative and sets them to zero. This automatically reduces the
    #   total error in the forecast and to compensate, I'm going to manually
    #   adjust the standard deviation so that it hits the target value
    adj_factor = 3.25
    adj_normalized_st_dev = normalized_std_dev * adj_factor
    logger.info(f"Adjusted normalized standard deviation"
                f": {adj_normalized_st_dev}")

    def _add_error(capacity, profile, dso_meta, idx):
        st_dev_profile = [x * adj_normalized_st_dev for x in profile]
        normalized_error = np.random.normal(0, st_dev_profile, (len(profile)))
        error = normalized_error  # * capacity
        forecast_profile = [x + y for x, y in zip(profile, error)]
        # error = error.tolist()
        forecast_profile = forecast_cleanup(dso_meta, idx, error, profile,
                                            forecast_profile)

        return forecast_profile

    for idx in dso_idxs:
        if idx < 200:
            logger.info('Creating utility solar forecast profile for DSO {}...'.format(dso_meta[idx]['200-bus']))
            profile = _add_extra_days_to_hourly(dso_meta[idx]['utility power profile'])
            dso_meta[idx]['utility power profile'] = profile
            forecast_profile = _add_error(
                dso_meta[idx]['utility solar capacity'],
                profile,
                dso_meta,
                idx)
            dso_meta[idx]['utility forecast profile'] = forecast_profile

            filename = 'DSO_{}_utility_hourly_forecast_power_profile.csv'.format(dso_meta[idx]['200-bus'])
            head, tail = os.path.split(dso_meta[idx]['utility power profile path'])
            outpath = os.path.join(head, filename)
            write_power_profile(outpath, forecast_profile)
            logger.info('\t...and saving output forecast file to {}'.format(outpath))

            logger.info('Creating distributed solar forecast profile for DSO {}'.format(dso_meta[idx]['200-bus']))
            profile = _add_extra_days_to_hourly(dso_meta[idx]['dist power profile'])
            dso_meta[idx]['dist power profile'] = profile
            forecast_profile = _add_error(
                dso_meta[idx]['dist solar capacity'],
                profile,
                dso_meta,
                idx)
            dso_meta[idx]['dist forecast profile'] = forecast_profile
            filename = 'DSO_{}_dist_hourly_forecast_power_profile.csv'.format(dso_meta[idx]['200-bus'])
            head, tail = os.path.split(dso_meta[idx]['dist power profile path'])
            outpath = os.path.join(head, filename)
            write_power_profile(outpath, forecast_profile)
            logger.info('\t...and saving output forecast file to {}'.format(outpath))
        else:
            logger.info('Creating utility solar forecast profile for DSO {}...'.format(dso_meta[idx]['8-node dso num']))
            profile = _add_extra_days_to_hourly(dso_meta[idx]['utility power profile'])
            dso_meta[idx]['utility power profile'] = profile
            forecast_profile = _add_error(
                dso_meta[idx]['utility solar capacity'],
                profile,
                dso_meta,
                idx)
            dso_meta[idx]['utility forecast profile'] = forecast_profile
            filename = '8-node DSO_{}_utility_hourly_forecast_power_profile.csv'.format(dso_meta[idx]['8-node dso num'])
            head, tail = os.path.split(dso_meta[idx]['utility power profile path'])
            outpath = os.path.join(head, filename)
            write_power_profile(outpath, forecast_profile)
            logger.info('\t...and saving output forecast file to {}'.format(outpath))

            logger.info(
                'Creating distributed solar forecast profile for DSO {}'.format(dso_meta[idx]['8-node dso num']))
            profile = _add_extra_days_to_hourly(dso_meta[idx]['dist power profile'])
            dso_meta[idx]['dist power profile'] = profile
            forecast_profile = _add_error(
                dso_meta[idx]['dist solar capacity'],
                profile,
                dso_meta,
                idx)
            dso_meta[idx]['dist forecast profile'] = forecast_profile

    # Creating single file for all distributed DSO solar PV forecasts
    if dso_type == 200:
        filename = '200-node_dist_hourly_forecast_power.csv'
        head, tail = os.path.split(dso_meta[idx]['dist power profile path'])
    else:
        filename = '8-node_dist_hourly_forecast_power.csv'
        head, tail = os.path.split(dso_meta[idx]['dist power profile path'])
    head, tail = os.path.split(head)
    outpath = os.path.join(head, filename)
    dso_fh = _open_file(outpath, 'w')
    ts = dt.datetime(2015, 12, 29, 0, 0, 0)
    # For some reason, datetime deltas don't work in hours
    # Specifying the increment in terms of seconds (which are supported).
    time_inc = dt.timedelta(seconds=3600)

    # Outermost loop is the timestep of the power profile
    # Arbitrarily picking first profile to determine the length of
    #   the profile. All profiles should be the same length.
    if dso_type == 200:
        ts_idx_max = len(dso_meta[0]['dist forecast profile'])
    else:
        ts_idx_max = len(dso_meta[200]['dist forecast profile'])
    for ts_idx in range(0, ts_idx_max):
        out_str = ts.strftime('%Y-%m-%d %H:%M:%S')
        out_str = out_str + ','
        for idx in dso_idxs:
            out_str = out_str + str(
                truncate(dso_meta[idx]['dist forecast profile'][ts_idx]))
            if dso_type == 200:
                if idx == 199:
                    out_str = out_str + '\n'
                else:
                    out_str = out_str + ','
            else:
                if idx == 207:
                    out_str = out_str + '\n'
                else:
                    out_str = out_str + ','
        dso_fh.write(out_str)
        ts = ts + time_inc
    dso_fh.close()

    # Creating single forecast file in 200-node format. This only needs
    #   to be done if we're creating the 200-node version.
    if dso_type == 200:
        filename = 'DSOT utility hourly MW solar power forecast.csv'

        # Finding path based on previously generated files. The path I need
        #   is one level up from where the DSO-specific data is written
        head, tail = os.path.split(dso_meta[0]['utility power profile path'])
        head, tail = os.path.split(head)
        all_dso_fh = _create_200_node_csv_file_and_headers(filename, head)

        # Setting up timestamp for the file
        ts = dt.datetime(2015, 12, 29, 0, 0, 0)

        # For some reason, datetime deltas don't work in hours
        #   specifying the increment in terms of seconds (which are supported).
        time_inc = dt.timedelta(seconds=3600)

        # Augmenting hourly profile values with extra buffer days before
        #   January 1 and Leap Day. Saving the profile data, so I don't
        #   have to pull it all out again.
        profile_list = []
        for idx in range(0, 200):
            profile = _add_extra_days_to_hourly(
                dso_meta[idx]['utility forecast profile'])
            profile_list.append(profile)

        # Outermost loop is the timestep of the power profile
        # Arbitrarily picking first profile to determine the length of
        #   the profile. All profiles should be the same length.
        ts_idx_max = len(profile_list[0])
        for ts_idx in range(0, ts_idx_max):
            out_str = ts.strftime('%Y-%m-%d %H:%M:%S')
            out_str = out_str + ','
            for idx in range(0, 200):
                out_str = out_str + str(truncate(profile_list[idx][ts_idx]))
                if idx == 199:
                    out_str = out_str + '\n'
                else:
                    out_str = out_str + ','
            all_dso_fh.write(out_str)
            ts = ts + time_inc
        all_dso_fh.close()
    return dso_meta


def create_GLD_files(dso_meta):
    """
    This function creates 5-minute power profiles from hourly
    profiles through the power of linear interpolation technology (TM).
    These profiles are formatted as GridLAB-D tape players.

    Example:
    2008-12-25 00:00:00,622368.3864
    2008-12-25 00:05:00,738071.2498
    2008-12-25 00:10:00,680611.3676
    2008-12-25 00:15:00,696280.9035


    It only acts on the distributed power profiles since these are the
    ones that are needed by the distribution system simulator to be used
    to implement distributed (rooftop) generation in the DSO.

    Args:
        dso_meta (list): List of dicts with metadata associated with
            each DSO including the current distributed power profiles.
    Returns:
        dso-Meta (list): List of dicts with DSO data including the
        location of the interpolated values. Decided not to save these
        inside the dictionary themselves due to their size and the fact
        that I don't anticipate needing to reuse them.
    """
    samples_per_hour = 12
    for idx, dso in enumerate(dso_meta):
        if idx < 200:
            logger.info('Interpolating profile on DSO {}...'.format(
                dso['200-bus']))
        else:
            logger.info('Interpolating profile on 8-node DSO {}...'.format(
                dso['8-node dso num']))
        profile = dso['dist power profile']

        profile = _add_extra_days_to_hourly(profile)

        # Performing the actual interpolation
        #   x = the time intervals over which the interpolation occurs
        #   y = the data points bounding the interpolation
        #   xp = the new times at which interpolation occurs
        length = len(profile)
        x = np.linspace(0, length, num=length, endpoint=False)
        y = profile
        xp = np.linspace(0, length, length * samples_per_hour)
        hf_profile = np.interp(xp, x, y)

        # Generating GLD-formatted timestamps
        # Setting up timestamp for the file
        ts = dt.datetime(2015, 12, 29, 0, 0, 0)
        # For some reason, datetime deltas don't work in hours
        #   specifying the increment in terms of seconds (which are supported).
        time_inc = dt.timedelta(seconds=300)

        # Getting the file open and ready for writing...
        head, tail = os.path.split(dso['dist power profile path'])
        if idx < 200:
            filename = 'DSO_{}_5_minute_dist_power.csv'.format(
                dso['200-bus'])
        else:
            filename = 'DSO_{}_5_minute_dist_power.csv'.format(
                dso['8-node dso num'])
        outpath = os.path.join(head, filename)
        dso['5 min dist power profile path'] = outpath
        gld_fh = _open_file(outpath, 'w')

        length = len(hf_profile)
        for ts_idx in range(0, length):
            out_str = ts.strftime('%Y-%m-%d %H:%M:%S')
            out_str = out_str + ',' + str(truncate(hf_profile[ts_idx])) \
                      + '\n'
            gld_fh.write(out_str)
            ts = ts + time_inc
        gld_fh.close()

        dso_meta[idx] = dso
        logger.info('\t...and saved interpolated file at {}'.format(outpath))
    return dso_meta


def create_dsot_utility_solar_file(dso_meta, output_path):
    """
    This function creates a time-series "tape" file for ingest by
    the DSO+T co-simulation for each of the 200 utility-scale solar PV
    generation sites. The format is specific to this study (though not
    opaque in the least); a sample is shown below taken from the wind
    profile:

    time,wind26,wind28,wind29,...
    2015-12-29 00:00:00,238.100,646.600,365.400,...
    2015-12-29 01:00:00,231.497,646.600,365.400,...
    ...

    The data needed for these profiles starts prior to Jan 1 to allow the
    models to warm up. The data prior to Jan 1 is not crucial, and I'll just
    the Jan 1 data for those early days as a good-enough approximation.

    Also creates an interpolated 5-min version of the file for the
    RT market.

    Args:
        dso_meta (list): List of dicts with metadata associated with
            each DSO including the utility-scale solar PV power profiles.
        output_path (str): Path to directory where file will be written
    """

    dsot_fh = _create_200_node_csv_file_and_headers(
        'DSOT utility hourly MW solar power profile.csv',
        output_path)
    dsot_fm_fh = _create_200_node_csv_file_and_headers(
        'DSOT utility 5-min MW solar power profile.csv',
        output_path)

    num_dso = 200

    # Setting up timestamp for the file
    ts = dt.datetime(2015, 12, 29, 0, 0, 0)

    # For some reason, datetime deltas don't work in hours
    #   specifying the increment in terms of seconds (which are supported).
    time_inc = dt.timedelta(seconds=3600)
    time_inc_fm = dt.timedelta(seconds=300)

    # Setting up counters for the loop.
    hours_per_day = 24
    days_per_year = 366
    buffer_days = 3  # Days added to the beginning of the profile to warm up
    # the model

    max_idx = num_dso - 1
    max_ts_idx = (days_per_year + buffer_days) * hours_per_day
    for ts_idx in range(0, max_ts_idx):
        # Data structure for holding the five-minute interpolated values
        fm_data = []

        out_str = ts.strftime('%Y-%m-%d %H:%M:%S')
        out_str = out_str + ','

        for idx, dso in enumerate(dso_meta):
            out_str = out_str + str(truncate(dso['utility power profile'][
                                                 ts_idx])) + ','

            # Creating the 5-minute real-time tape in situ
            #
            # Since we're incrementing by DSO I need to build up the 5-min
            #   data as a chunk across DSOs and then write it out in
            #   hourly blocks
            # I'm working with the current and previous time steps,
            #   so I don't need to do anything on the first timestep
            #   (ts_idx = 0)
            # Don't need to create this file with the data from the
            #   aggregated 8-node DSOs (idx = 200..207)
            if ts_idx == 0:
                fm_data.append([])
                ts_fm = ts
                # Creates the interpolated data set leaving the last
                #   endpoint out (this will be the first endpoint in the
                #   next set).
                five_min_data = _five_minute_interpolation(
                    dso['utility power profile'][ts_idx],
                    dso['utility power profile'][ts_idx + 1])
                data_length = len(five_min_data)
            elif ts_idx > 0 and idx <= max_idx:
                fm_data.append([])
                ts_fm = ts
                # Creates the interpolated data set leaving the last
                #   endpoint out (this will be the first endpoint in the
                #   next set).
                five_min_data = _five_minute_interpolation(
                    dso['utility power profile'][ts_idx - 1],
                    dso['utility power profile'][ts_idx])
                data_length = len(five_min_data)

            for idx_fm, data in enumerate(five_min_data):
                # Create the extra timestamps for the first DSO only.
                if idx == 0:
                    date_str_fm = ts_fm.strftime('%Y-%m-%d %H:%M:%S')
                    ts_fm = ts_fm + time_inc_fm
                    data_str = date_str_fm + ',' + str(truncate(data)) + ','
                    fm_data[idx].append(data_str)
                elif idx <= max_idx - 1:
                    fm_data[idx].append(str(truncate(data)) + ',')
                elif idx == max_idx:
                    fm_data[idx].append(str(truncate(data)) + '\n')

            # Write out once we've interpolated data for all 200 DSOs.
            if idx == max_idx:
                out_str_fm = ''
                for time in range(0, data_length):
                    for dso_idx in range(0, num_dso):
                        out_str_fm = out_str_fm + fm_data[dso_idx][time]
                # Multi-line write to file
                dsot_fm_fh.write(out_str_fm)
                dsot_fm_fh.flush()

        # Removing last comma and replacing with a newline
        out_str = out_str[:-1]
        out_str = out_str + '\n'
        dsot_fh.write(out_str)
        ts = ts + time_inc

    dsot_fh.close()
    dsot_fm_fh.close()


def _five_minute_interpolation(data_1, data_2):
    # Performing the actual interpolation
    #   x = the time intervals over which the interpolation occurs
    #   y = the data points bounding the interpolation
    #   xp = the new times at which interpolation occurs
    x = [0, 1]
    y = [data_1, data_2]
    xp = np.linspace(0, 1, 12)
    data_12 = np.interp(xp, x, y)
    return data_12


def log_metdata(dso_meta):
    metadata_fields = ['200-bus', '8-bus', 'lat', 'long', 'avg load',
                       'load_fraction', 'total solar capacity',
                       'dist solar capacity', 'utility solar capacity']
    for dso in dso_meta:
        logging.info(f'DSO {dso["200-bus"]} metadata')
        for key in metadata_fields:
            if key in dso:
                logging.info(f'\t{key}: {dso[key]}')


def generate_forecast_metrics(dso_meta, output_path):
    """
    Calculates the RMSE associated with each solar forecast file and write it out to file.

    Args:
        dso_meta:
        output_path:
    Returns:
        (none)
    """

    output_file = os.path.join(output_path, 'Solar error calculations.csv')
    error_fh = _open_file(output_file, 'w')
    error_fh.write('DSO index,solar type,Capacity (MW),'
                   'Normalized RMSE (MW),Normalized StDev,Correlation '
                   'Coefficient\n')

    for idx, dso in enumerate(dso_meta):
        error_list = []
        error_list.append([x[1] - x[0] for x in
                           zip(dso['utility power profile'],
                               dso['utility forecast profile'])])
        error_list.append([x[1] - x[0] for x in
                           zip(dso['dist power profile'],
                               dso['dist forecast profile'])])
        solar_type = ['utility', 'distributed']
        for idx2, errors in enumerate(error_list):
            if idx2 == 0:
                capacity = dso['utility solar capacity']
            else:
                capacity = dso['dist solar capacity']

            # MAPE calculation - not possible without removing zero
            #   values from actual power prediction
            # abs_errors = [abs(x) for x in errors]
            # abs_error_factor = [x[0]/x[1] for x in zip(abs_errors,
            #                                          dso['utility power
            #                                          profile'])]
            # mape = sum(abs_error_factor) / len(abs_error_factor)

            # Correlation coefficient
            corr = np.corrcoef(dso['utility power profile'], dso['utility forecast profile'])

            normalized_errors = [x / capacity for x in errors]
            normalized_rmse = np.sqrt(np.mean([i ** 2 for i in normalized_errors]))
            normalized_stdev = np.std(normalized_errors)
            error_fh.write(f"{idx},{solar_type[idx2]},{capacity},"
                           f"{normalized_rmse},{normalized_stdev},"
                           f"{corr[0][1]}\n")
    error_fh.close()


def _auto_run(args):
    """
    This function executes when the script is called as a stand-alone
    executable.

    A more complete description of this code can be found in the
    docstring at the beginning of this file.

    Args:
        '-a' or '--auto_run_dir' - Path of the auto_run folder
        that contains examples of the files used in this script. Used
        for development purposes as well as models/examples/documentation
        of the format and contents expected in said files

        '-s' or '--solar_metadata' - Path to JSON file with all non-DSO
        metadata

        '-d' or '--dso_metadata' - Path to .xlsx file with all DSO
        metadata, file is part of the DSO+T planning/documentation
        dataset and should not need to be manually created or edited.

        '-o' or '--output_path' - Location to save the solar files

    Returns:
        (none)
    """

    # Must parse solar metadata file first as it contains the name of the
    #   Excel worksheet that contains the values for the DSO metadata.
    solar_meta = parse_solar_metadata(args.solar_metadata)
    # dso_meta = parse_DSO_metadata_Excel(args.dso_metadata,
    #                              solar_meta['dso_metadata_worksheet_name'])
    dso_meta = parse_DSO_metadata_Excel_JSON(args.dso_metadata_Excel,
                                             solar_meta[
                                                 'dso_metadata_worksheet_name'],
                                             args.dso_metadata_JSON)
    build_dso_solar_folders(dso_meta, args.nsrdb_output_path)
    dso_meta = add_locations(dso_meta, solar_meta, args.nsrdb_output_path)
    dso_meta = download_nsrdb_data(dso_meta, solar_meta,
                                   args.nsrdb_output_path)
    # generate_KML(dso_meta, args.csv_kml_output)
    build_dso_solar_folders(dso_meta, args.solar_pv_power_output_path)
    dso_meta = calc_solarPV_power(dso_meta, args.solar_pv_power_output_path)
    dso_meta = calc_dso_solar_fraction(dso_meta)
    dso_meta = aggregate_scale_solar_pv_profiles(dso_meta, solar_meta,
                                                 args.solar_pv_power_output_path)
    dso_meta = aggregate_to_8_nodes(dso_meta, args.solar_pv_power_output_path)
    dso_meta = create_hourly_solar_forecast(dso_meta, 200, args.random_seed)
    dso_meta = create_hourly_solar_forecast(dso_meta, 8, args.random_seed)
    generate_forecast_metrics(dso_meta, args.solar_pv_power_output_path)
    dso_meta = create_GLD_files(dso_meta)
    create_dsot_utility_solar_file(dso_meta, args.solar_pv_power_output_path)

    # Graphs are re-created every run and can take several minutes
    # create_graphs(dso_meta, 'profiles')
    # create_graphs(dso_meta, 'utility forecast')


if __name__ == '__main__':
    # TDH: This slightly complex mess allows lower importance messages
    # to be sent to the log file and ERROR messages to additionally
    # be sent to the console as well. Thus, when bad things happen
    # the user will get an error message in both places which,
    # hopefully, will aid in troubleshooting.
    fileHandle = logging.FileHandler("dsot_solar.log", mode='w')
    fileHandle.setLevel(logging.DEBUG)
    streamHandle = logging.StreamHandler(sys.stdout)
    streamHandle.setLevel(logging.ERROR)
    logging.basicConfig(level=logging.INFO,
                        handlers=[fileHandle, streamHandle])

    parser = argparse.ArgumentParser(description='Download NSRDB data.')
    # TDH: Have to do a bit of work to generate a good default
    # path for the auto_run folder where the development test data is
    # held.
    script_path = os.path.dirname(os.path.realpath(__file__))
    auto_run_dir = os.path.join(script_path, 'auto_run')
    parser.add_argument('-a',
                        '--auto_run_dir',
                        nargs='?',
                        default=auto_run_dir)
    script_path = Path(script_path)
    tesp_root = script_path.parents[3]
    dsot_data_path = os.path.join(tesp_root, 'examples', 'data')
    solar_file = os.path.join(dsot_data_path, 'dsot_solar_metadata.json')
    parser.add_argument('-s',
                        '--solar_metadata',
                        nargs='?',
                        default=solar_file)

    csv_kml_file = os.path.join(dsot_data_path, 'dsot_solar.kml')
    parser.add_argument('-k',
                        '--csv_kml_output',
                        nargs='?',
                        default=csv_kml_file)

    dso_file_Excel = os.path.join(dsot_data_path, 'bus_generators.xlsx')
    parser.add_argument('-x',
                        '--dso_metadata_Excel',
                        nargs='?',
                        default=dso_file_Excel)

    dso_file_JSON = os.path.join(dsot_data_path, '200-hi-metadata-lean.json')
    parser.add_argument('-j',
                        '--dso_metadata_JSON',
                        nargs='?',
                        default=dso_file_JSON)

    nsrdb_output_path = os.path.join(dsot_data_path,
                                     'solar_data',
                                     'nsrdb_files')
    parser.add_argument('-n',
                        '--nsrdb_output_path',
                        nargs='?',
                        default=nsrdb_output_path)

    solar_pv_power_output_path = os.path.join(dsot_data_path,
                                              'solar_data',
                                              'solar_pv_power_profiles')
    parser.add_argument('-p',
                        '--solar_pv_power_output_path',
                        nargs='?',
                        default=solar_pv_power_output_path)

    parser.add_argument('-r',
                        '--random_seed',
                        nargs='?',
                        default=0)

    args = parser.parse_args()
    _auto_run(args)
